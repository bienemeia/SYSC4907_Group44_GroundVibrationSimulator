import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication

import sqlite3
import openpyxl

def convert_db_to_excel(db_file, excel_file):
    # Connect to the database
    conn = sqlite3.connect(db_file)
    cursor = conn.cursor()

    # Execute the SQL query to fetch data
    query = "SELECT * FROM EXPERIMENT"
    cursor.execute(query)

    # Fetch all the rows of the result
    rows = cursor.fetchall()

    # Create a new Excel workbook
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Write the header row of the Excel sheet
    header_row = [description[0] for description in cursor.description]
    for i, header in enumerate(header_row):
        worksheet.cell(row=1, column=i + 1, value=header)

    # Write the data rows to the Excel sheet
    for i, row in enumerate(rows):
        for j, value in enumerate(row):
            worksheet.cell(row=i + 2, column=j + 1, value=value)

    # Save the workbook to an Excel file
    workbook.save(excel_file)

    # Close the database connection
    conn.close()


def send_email_with_attachment(to, subject, body, file_path):
    msg = MIMEMultipart()
    msg['From'] = 'turtleprojecttest@gmail.com'
    msg['To'] = to
    msg['Subject'] = subject

    msg.attach(MIMEText(body))
    
    with open(file_path, "rb") as f:
        attach = MIMEApplication(f.read(),_subtype="py")
        attach.add_header('content-disposition', 'attachment', filename=file_path)
        msg.attach(attach)

    try:
        smtp_server = smtplib.SMTP('smtp.gmail.com', 587)
        smtp_server.ehlo()
        smtp_server.starttls()
        smtp_server.ehlo()
        smtp_server.login('turtleprojecttest@gmail.com', 'loeqkblwfwvnbbtd')
        smtp_server.sendmail('turtleprojecttest@gmail.com', to, msg.as_string())
        print("Email sent successfully")
    except Exception as e:
        print("Error: ", e)
    finally:
        smtp_server.quit()


def main():
	convert_db_to_excel("OralDemo.db", "database.xlsx")
	send_email_with_attachment("shawaizkhan.c@gmail.com","database","here","/home/pi/Desktop/database.xlsx")

if __name__ == "__main__":
	main()
